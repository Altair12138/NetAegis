"""Excel 报表导出。

Sheet 设计：
- Overview       — Job 概览（设备总数、成功/失败/命名不一致计数、耗时）
- Devices        — 每台设备一行（设备名/IP/厂商/类型/状态/耗时/log/json）
- Errors         — 失败明细（设备、命令、错误信息）
- NameMismatch   — 命名不一致清单（设备、IP、实际 hostname）
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from loguru import logger

from .db import DeviceRunRow, JobRow, session


_HEADER_FILL = PatternFill("solid", fgColor="305496")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_BAD_FILL = PatternFill("solid", fgColor="F8CBAD")
_WARN_FILL = PatternFill("solid", fgColor="FFE699")


def generate(job_id: str, out_path: Path | None = None) -> Path:
    with session() as s:
        job = s.get(JobRow, job_id)
        if not job:
            raise KeyError(f"job not found: {job_id}")
        devices = list(s.query(DeviceRunRow).filter(DeviceRunRow.job_id == job_id).all())

    wb = Workbook()

    _sheet_overview(wb, job, devices)
    _sheet_devices(wb, devices)
    _sheet_errors(wb, devices)
    _sheet_name_mismatch(wb, devices)

    if out_path is None:
        out_path = Path(job.result_dir) / f"report_{job_id}.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------

def _sheet_overview(wb: Workbook, job: JobRow, devices: list[DeviceRunRow]) -> None:
    ws = wb.active
    ws.title = "Overview"

    counts: dict[str, int] = {}
    for d in devices:
        counts[d.status] = counts.get(d.status, 0) + 1
    total = len(devices)
    duration = ((job.finished_at or datetime.now()) - (job.started_at or job.created_at)).total_seconds()

    rows = [
        ("Job ID", job.id),
        ("Type", job.type),
        ("Status", job.status),
        ("Concurrency", job.concurrency),
        ("Created", job.created_at),
        ("Started", job.started_at),
        ("Finished", job.finished_at),
        ("Duration (s)", round(duration, 1)),
        ("Result dir", job.result_dir),
        ("", ""),
        ("Total devices", total),
        ("Success", counts.get("success", 0)),
        ("Name mismatch", counts.get("name_mismatch", 0)),
        ("Failed", counts.get("failed", 0)),
        ("Skipped (canceled)", counts.get("skipped", 0)),
    ]
    for r, (k, v) in enumerate(rows, start=1):
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60


def _sheet_devices(wb: Workbook, devices: list[DeviceRunRow]) -> None:
    ws = wb.create_sheet("Devices")
    headers = ["Device", "Mgmt IP", "Status", "NameMismatch",
               "Started", "Finished", "Duration(s)", "Log", "JSON", "Error"]
    _write_header(ws, headers)
    for i, d in enumerate(devices, start=2):
        dur = ""
        if d.started_at and d.finished_at:
            dur = round((d.finished_at - d.started_at).total_seconds(), 1)
        row = [d.device_name, d.mgmt_ip, d.status, "Y" if d.name_mismatch else "",
               d.started_at, d.finished_at, dur,
               d.log_path or "", d.json_path or "", d.error or ""]
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if d.status == "failed":
                cell.fill = _BAD_FILL
            elif d.status == "name_mismatch":
                cell.fill = _WARN_FILL
    _autosize(ws, headers)


def _sheet_errors(wb: Workbook, devices: list[DeviceRunRow]) -> None:
    ws = wb.create_sheet("Errors")
    headers = ["Device", "Mgmt IP", "Command Key", "Error"]
    _write_header(ws, headers)
    row_i = 2
    for d in devices:
        # 优先从同目录 JSON 提取每条命令的 error；db 里只有汇总 error
        details = _load_command_errors(d.json_path)
        if not details and d.error:
            ws.cell(row=row_i, column=1, value=d.device_name)
            ws.cell(row=row_i, column=2, value=d.mgmt_ip)
            ws.cell(row=row_i, column=3, value="-")
            ws.cell(row=row_i, column=4, value=d.error)
            row_i += 1
            continue
        for key, err in details:
            ws.cell(row=row_i, column=1, value=d.device_name)
            ws.cell(row=row_i, column=2, value=d.mgmt_ip)
            ws.cell(row=row_i, column=3, value=key)
            ws.cell(row=row_i, column=4, value=err)
            row_i += 1
    _autosize(ws, headers)


def _sheet_name_mismatch(wb: Workbook, devices: list[DeviceRunRow]) -> None:
    ws = wb.create_sheet("NameMismatch")
    headers = ["Device (expected)", "Mgmt IP", "Actual hostname", "JSON"]
    _write_header(ws, headers)
    row_i = 2
    for d in devices:
        if not d.name_mismatch:
            continue
        actual = ""
        if d.json_path and Path(d.json_path).exists():
            try:
                data = json.loads(Path(d.json_path).read_text(encoding="utf-8"))
                actual = data.get("name_check", {}).get("actual", "") or ""
            except Exception:  # noqa: BLE001
                logger.warning(f"Failed to parse JSON for name mismatch: {d.json_path}")
                actual = ""
        ws.cell(row=row_i, column=1, value=d.device_name)
        ws.cell(row=row_i, column=2, value=d.mgmt_ip)
        ws.cell(row=row_i, column=3, value=actual)
        ws.cell(row=row_i, column=4, value=d.json_path or "")
        row_i += 1
    _autosize(ws, headers)


# ---------------------------------------------------------------------------

def _write_header(ws, headers: list[str]) -> None:
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _autosize(ws, headers: list[str]) -> None:
    for col_idx in range(1, len(headers) + 1):
        letter = get_column_letter(col_idx)
        max_len = len(headers[col_idx - 1])
        for cell in ws[letter]:
            v = cell.value
            if v is None:
                continue
            max_len = max(max_len, min(60, len(str(v))))
        ws.column_dimensions[letter].width = max_len + 2


def _load_command_errors(json_path: str | None) -> list[tuple[str, str]]:
    if not json_path or not Path(json_path).exists():
        return []
    try:
        data = json.loads(Path(json_path).read_text(encoding="utf-8"))
    except Exception:  # noqa: BLE001
        logger.warning(f"Failed to load command errors from {json_path}")
        return []
    out = []
    for c in data.get("commands", []):
        if c.get("error"):
            out.append((c.get("key", "-"), str(c["error"])))
    return out
